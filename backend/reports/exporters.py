from __future__ import annotations

import json
from io import BytesIO
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas


def to_xlsx_table(rows: List[Dict[str, Any]], title: str = "Report") -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = []
    if rows:
        # Preserve key order from the first row
        headers = list(rows[0].keys())
    else:
        headers = ["No data"]

    ws.append(headers)
    for r in rows:
        row_vals = []
        for h in headers:
            v = r.get(h, "")
            if isinstance(v, (list, dict)):
                v = json.dumps(v, default=str)
            elif v is None:
                v = ""
            row_vals.append(v)
        ws.append(row_vals)

    # Basic auto-width
    for idx, h in enumerate(headers, start=1):
        col = get_column_letter(idx)
        max_len = len(str(h))
        for cell in ws[col]:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col].width = min(60, max(12, max_len + 2))

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _draw_wrapped_text(c: canvas.Canvas, x: int, y: int, text: str, max_width: int, line_height: int = 14) -> int:
    """Draw text with rough wrapping; returns updated y."""
    words = (text or "").split()
    if not words:
        c.drawString(x, y, "")
        return y - line_height

    line = ""
    for w in words:
        test = (line + " " + w).strip()
        if c.stringWidth(test, "Helvetica", 10) <= max_width:
            line = test
        else:
            c.drawString(x, y, line)
            y -= line_height
            line = w
    if line:
        c.drawString(x, y, line)
        y -= line_height
    return y


def to_pdf_summary(data: Dict[str, Any], title: str = "Report") -> BytesIO:
    out = BytesIO()
    c = canvas.Canvas(out, pagesize=letter)
    width, height = letter

    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, height - 50, title)

    c.setFont("Helvetica", 10)
    y = height - 80

    # Flatten dict into key/value lines (with sections)
    def render(obj, indent=0):
        nonlocal y
        if y < 80:
            c.showPage()
            c.setFont("Helvetica", 10)
            y = height - 50
        if isinstance(obj, dict):
            for k, v in obj.items():
                if isinstance(v, (dict, list)):
                    c.setFont("Helvetica-Bold", 10)
                    c.drawString(50 + indent, y, str(k))
                    y -= 14
                    c.setFont("Helvetica", 10)
                    render(v, indent=indent + 12)
                else:
                    line = f"{k}: {v}"
                    y = _draw_wrapped_text(c, 50 + indent, y, line, max_width=int(width - 100 - indent))
        elif isinstance(obj, list):
            for item in obj:
                render(item, indent=indent)
                y -= 6
        else:
            y = _draw_wrapped_text(c, 50 + indent, y, str(obj), max_width=int(width - 100 - indent))

    render(data)

    c.showPage()
    c.save()
    out.seek(0)
    return out


def to_pdf_table(rows: List[Dict[str, Any]], title: str = "Report") -> BytesIO:
    out = BytesIO()
    c = canvas.Canvas(out, pagesize=letter)
    width, height = letter

    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, height - 50, title)

    y = height - 80
    c.setFont("Helvetica", 9)

    if not rows:
        c.drawString(50, y, "No data")
        c.showPage()
        c.save()
        out.seek(0)
        return out

    headers = list(rows[0].keys())
    col_count = len(headers)
    col_width = max(60, int((width - 100) / max(1, col_count)))

    def new_page():
        nonlocal y
        c.showPage()
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, height - 50, title)
        c.setFont("Helvetica", 9)
        y = height - 80

    # header
    c.setFont("Helvetica-Bold", 9)
    x = 50
    for h in headers:
        c.drawString(x, y, str(h)[:30])
        x += col_width
    c.setFont("Helvetica", 9)
    y -= 14

    for r in rows:
        if y < 60:
            new_page()
            c.setFont("Helvetica-Bold", 9)
            x = 50
            for h in headers:
                c.drawString(x, y, str(h)[:30])
                x += col_width
            c.setFont("Helvetica", 9)
            y -= 14

        x = 50
        for h in headers:
            val = r.get(h, "")
            c.drawString(x, y, str(val)[:30])
            x += col_width
        y -= 12

    c.showPage()
    c.save()
    out.seek(0)
    return out
